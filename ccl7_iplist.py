import openpyxl
import ChargingContextAai
import time

def getServiceListByList(sheet,startRow):
    changeLag_col = 3
    serviceId_col = 8
    serviceName_col = 10
    ipAddressL3_col = 13
    protocolNumber_col = 14
    portNumberL4_col = 15
    urlL7_col = 16
    #firstLineServiceId = sheet.cell(row=startRow, column=serviceId_col).value
    retList = []

    for rowNumber in range(startRow,sheet.max_row + 1):
        changeLag = sheet.cell(row=rowNumber, column=changeLag_col).value
        serviceId = sheet.cell(row=rowNumber, column=serviceId_col).value
        serviceName = sheet.cell(row=rowNumber, column=serviceName_col).value
        ipAddressL3 = sheet.cell(row=rowNumber, column=ipAddressL3_col).value
        protocolNumber = sheet.cell(row=rowNumber, column=protocolNumber_col).value
        portNumberL4 = sheet.cell(row=rowNumber, column=portNumberL4_col).value
        if portNumberL4 != None:
            portNumberL4 = str(portNumberL4)
        urlL7 = sheet.cell(row=rowNumber, column=urlL7_col).value

        layerLag = "L7"
        portsplit_1 = "|"
        portsplit_2 = "=="
        portL4List = []
        portstr = portNumberL4
        if portNumberL4 != None:
            if portsplit_2 in str(portNumberL4):
                portstr = portNumberL4.split(portsplit_2)[0]
                portL4List.append(portNumberL4.split(portsplit_2)[1])
            else:
                pass
            if portsplit_1 in str(portstr):
                portstr = portstr.split(portsplit_1)
                for p in portstr:
                    portL4List.append(p)
            else:
                portL4List.append(portstr)
            for p in portL4List:
                retList.append((layerLag, changeLag, serviceId, serviceName, ipAddressL3, protocolNumber, p, urlL7))
        else:
            retList.append((layerLag, changeLag, serviceId, serviceName, ipAddressL3, protocolNumber, portNumberL4, urlL7))
        if layerLag == "L7" and serviceName == None:
            pass
    retList = list(set(retList))

    return  retList

def arrangeTheList(lst):

    sList = []
    retList = []
    tempList = []
    for tup in lst:
        layerLag, changeLag, serviceId, serviceName, ipAddress, protocolNumber, portNumber, url = tup
        sList.append(serviceId)
    sList = list(set(sList))
    for sValue in sList:
        for tup in lst:
            layerLag, changeLag, serviceId, serviceName, ipAddress, protocolNumber, portNumber, url = tup
            if sValue == serviceId:
                tempList.append(tup)
        retList.append(tempList)
        tempList = []
    return retList

def arrangeTheList_2(rtList):
    new_ret_list = []
    for lst in rtList:
        tmplist = []
        for tup in lst:
            if tup[1] == "新增":
                tmplist.append(tup)
        for tup in lst:
            if tup[1] == "删除":
                tmplist.append(tup)
        new_ret_list.append(tmplist)
    return new_ret_list




def addServiceUrlIpListToDict(lst):
    global serviceUrlIpListDict
    urlIpListDict = {}

    serviceName = lst[0][3]

    urlList = []
    for tup in lst:
        urlporttup = (tup[7],tup[6])
        urlList.append(urlporttup)
    urlList = list(set(urlList))
    for url in urlList:
        urlIpListDict[url] = []
    serviceUrlIpListDict[serviceName] = urlIpListDict
    for urlKey in serviceUrlIpListDict[serviceName]:
        ipPrefixList = []
        ipPrefixList = getIpPrefixListByUrl_serviceName(urlKey,serviceName)
        urlIpListDict[urlKey] = ipPrefixList



def getIpPrefixListByUrl_serviceName(urlport,service_name):
    global configList
    retList = []
    #print("url port servicename",urlport,service_name)
    url,service_port = urlport
    http_host,http_uri,port = ChargingContextAai.processUrl(url)
    #print("host+++",url,service_name,http_host,http_uri,port,service_port)

    for i in range(0,len(configList)):
        if http_host in configList[i]:
            k = i
            tmplist = []
            for j in range(k,len(configList)):
                if "no shutdown" in configList[j]:
                    e = j
                    #print("++++++++++++",service_name,service_port,http_host,http_uri,port,ChargingContextAai.EntryIsTrue(service_name,service_port,http_host,http_uri,port,configList,k-1,e))
                    if ChargingContextAai.EntryIsTrue(service_name,service_port,http_host,http_uri,port,configList,k-1,e) ==True:

                        for t in range(k,e):
                            if 'server-address eq ip-prefix-list' in configList[t]:
                                if 'server-address eq ip-prefix-list "app_' + service_name in configList[t]:
                                    ipListStr = configList[t].split("server-address eq ")[1].replace("\n", "")
                                    #print(str(urlport)+"获取的iplst名："+ipListStr)
                                    tmplist = getTheIpPrefixList(ipListStr)
                        if len(tmplist) != 0:
                            retList.append(tmplist)
                    break
    '''
    http_host = 'expression 1 http-host eq "'+ http_host+'"'
    if http_url != None:
        http_url = 'expression 2 http-uri eq "'+ http_url+'"'

    for i in range(0,len(configList)):
        if http_host in configList[i]:
            #print("5555++",i,configList[i],configList[i+1])
            if http_url !=None and http_url in configList[i+1]:
                #print("66666+++",http_host,http_url)
                k = i
                tmplist = []
                for j in range(k,len(configList)):
                    if 'no shutdown' in configList[j]:
                        break
                    if 'server-address eq ip-prefix-list "app_'+service_name in configList[j]:
                        ipListStr = configList[j].split("server-address eq ")[1].replace("\n","")
                        #print(ipListStr)
                        tmplist = getTheIpPrefixList(ipListStr)
                if len(tmplist)!=0:
                    retList.append(tmplist)
            elif http_url ==None:
                #print("7777+++", http_host, http_url)
                k = i
                tmplist = []
                for j in range(k, len(configList)):
                    if 'no shutdown' in configList[j]:
                        break
                    if 'server-address eq ip-prefix-list "app_' + service_name in configList[j]:
                        ipListStr = configList[j].split("server-address eq ")[1].replace("\n", "")
                        # print(ipListStr)
                        tmplist = getTheIpPrefixList(ipListStr)
                if len(tmplist) != 0:
                    retList.append(tmplist)
    '''
    return retList

def getTheIpPrefixList(ip_list_str):
    global configList
    retList = []
    #print("++++",ip_list_str)
    retList.append(ip_list_str+" create")
    for i in range(0,len(configList)):
        if ip_list_str+' create' in configList[i]:
            f = i
            for j in range(f,len(configList)):
                if "exit" in configList[j]:
                    e = j
                    break
    for i in range(f,e):
        if 'prefix ' in configList[i]:
            #ipstr = configList[i].replace("\n","").split("prefix ")[1].split(" name")[0].replace("/32","")
            ipstr = configList[i].replace("\n", "").replace("    ", "")
            retList.append(ipstr)
    return retList

def setServicePostFixNum(service_name,postList):
    global configList
    for text in configList:
        if 'ip-prefix-list "app_'+service_name+'_' in text and "create" in text:
            postfixnumstr = text.split('ip-prefix-list "app_'+service_name+'_')[1].split('" create')[0]
            postList.append(int(postfixnumstr))


def addServiceUrlIpListToUserDict(lst):
    global serviceUrlIpListUserDict
    global del_serviceUrlIpListUserDict
    global servce_ipostfix_num
    serviceName = lst[0][3]

    if serviceName not in servce_ipostfix_num:
        servce_ipostfix_num[serviceName] = []
        setServicePostFixNum(serviceName,servce_ipostfix_num[serviceName])

    urlIpListDict = {}
    urlportList = []
    for tup in lst:
        urlport = (tup[7],tup[6])
        urlportList.append(urlport)
    urlportList = list(set(urlportList))
    for urlport in urlportList:
        urlIpListDict[urlport] = []

    for urlport in urlportList:
        #tmpdict = {}
        tmplst = []
        delTmplst = []
        for tup in lst:
            #print("addtup-----", tup,urlport)
            if tup[7] == urlport[0] and tup[6] == urlport[1] and tup[1]=="新增":
                #print("addtup++",tup)
                tmplst.append(tup[4])
            #if tup[7] == url and tup[1]=="删除":
             #   delTmplst.append(tup[4])
        urlIpListDict[urlport] = tmplst
        #del_urlIpListDict[url] = delTmplst
    serviceUrlIpListUserDict[serviceName] = urlIpListDict
    #del_serviceUrlIpListUserDict[serviceName] = del_urlIpListDict

def getTheCompatibleEntryIdByDict():
    global serviceEntryIdDict
    global allEntryIdList
    global allEntryIdDict
    serviceCaseStr = "no_head"
    retId = -1
    for i in range(20501,60000):
        if i not in allEntryIdDict[serviceCaseStr]:
            retId = i
            allEntryIdDict[serviceCaseStr].append(retId)
            break
    return retId



def  addCommandTocommandList(comlst,serverName, urlport,addList):
    global serviceUrlIpListDict
    global ip_prefix_list_max_number
    global servce_ipostfix_num

    ipstrDict = {}

    config_ipPrefixList = serviceUrlIpListDict[serverName][urlport]

    #若配置列表为空则表示配置文件中无该业务的ip_prefix_list

    if len(config_ipPrefixList) == 0:
        tl = []
        if len(servce_ipostfix_num[serverName]) == 0:
            postFixNum = 1
            servce_ipostfix_num[serverName].append(postFixNum)
        else:
            postFixNum = max(servce_ipostfix_num[serverName])+1
            servce_ipostfix_num[serverName].append(postFixNum)
        if postFixNum <10:
            ipliststr = 'ip-prefix-list "app_' + serverName + '_0' + str(postFixNum) + '"'
        else:
            ipliststr = 'ip-prefix-list "app_' + serverName + '_'+ str(postFixNum) + '"'
        tl.append(ipliststr)
        #entryId = getTheCompatibleEntryIdByDict()
        #createIpPrefixListEntry(comlst, serverName, url, ipliststr, entryId)
        config_ipPrefixList.append(tl)
        while len(addList) != 0:
            for lst in config_ipPrefixList:
                if len(lst) < ip_prefix_list_max_number + 1:
                    addIpStr = addList[0]
                    lst.append(addList[0])
                    #addIpPrefixListCommand(comlst, serverName, lst[0], addIpStr)
                    addList.remove(addList[0])
                    if len(addList) == 0:
                        break

            #若所有列表都满了则需创建新的列表
            if ip_prefix_list_is_full(config_ipPrefixList) == True:
                tl = []
                if len(servce_ipostfix_num[serverName]) != 0:
                    postFixNum = max(servce_ipostfix_num[serverName])+1
                    servce_ipostfix_num[serverName].append(postFixNum)
                if postFixNum < 10:
                    ipliststr = 'ip-prefix-list "app_' + serverName + '_0' + str(postFixNum) + '"'
                else:
                    ipliststr = 'ip-prefix-list "app_' + serverName + '_'+ str(postFixNum) + '"'
                tl.append(ipliststr)
                #entryId = getTheCompatibleEntryIdByDict()
                #createIpPrefixListEntry(comlst, serverName, url, ipliststr, entryId)
                config_ipPrefixList.append(tl)
    else:
        #循环直到用户需要添加iplist(addList)列表中的数据添加完才跳出循环
        while len(addList) != 0:
            for lst in config_ipPrefixList:
                if len(lst) < ip_prefix_list_max_number + 1:
                    addIpStr = addList[0]
                    lst.append(addList[0])
                    #addIpPrefixListCommand(comlst, serverName, lst[0], addIpStr)
                    addList.remove(addList[0])
                    if len(addList) == 0:
                        break

            #若所有列表都满了则需创建新的列表
            if ip_prefix_list_is_full(config_ipPrefixList) == True:
                tl = []
                if len(servce_ipostfix_num[serverName]) != 0:
                    postFixNum = max(servce_ipostfix_num[serverName])+1
                    servce_ipostfix_num[serverName].append(postFixNum)
                if postFixNum < 10:
                    ipliststr = 'ip-prefix-list "app_' + serverName + '_0' + str(postFixNum) + '"'
                else:
                    ipliststr = 'ip-prefix-list "app_' + serverName + '_'+ str(postFixNum) + '"'
                tl.append(ipliststr)
                #entryId = getTheCompatibleEntryIdByDict()
                #createIpPrefixListEntry(comlst,serverName,url,ipliststr,entryId)
                config_ipPrefixList.append(tl)

def createIpPrefixListEntry(clst,service_name,urlport,ip_list_str,enId):
    clst.append("exit all\n")
    clst.append("configure application-assurance group 1:1 policy\n")
    clst.append("app-filter\n")
    clst.append("entry " + str(enId) + " create\n")
    url = urlport[0]
    service_port = urlport[1]
    expression_1 = None
    expression_2 = None
    http_port = None
    if url != None:
        expression_1, expression_2, http_port = ChargingContextAai.processUrl(url)

    if expression_1 != None:
        clst.append('expression 1 http-host eq "' + expression_1 + '"\n')
    if expression_2 != None:
        clst.append('expression 2 http-uri eq "' + expression_2 + '"\n')

    clst.append("server-address eq "+ip_list_str + '\n')
    if service_port !=None:
        clst.append('server-port eq '+service_port+"\n")
    clst.append('application "APP_' + service_name + '"\n')
    clst.append("no shutdown\n")
    clst.append("exit\n")

    '''
    clst.append('exit all\n')
    clst.append('configure application-assurance group 1:1\n')
    clst.append(ip_list_str + "\n")
    '''
    clst.append("\n")



def addIpPrefixListCommand(clst,sName,ipStr):
    if "/" not in ipStr:
        ipStr = ipStr +"/32"
    clst.append("prefix "+ipStr+' name "'+sName+'"'+"\n")

def ip_prefix_list_is_full(cfg_ipPrefixList):
    global ip_prefix_list_max_number
    for lst in cfg_ipPrefixList:
        if len(lst) < ip_prefix_list_max_number + 1:
            return False
    return True

# def DeleteTheIpPrefixList(comlst,server_name,url,del_list):
#     pass


# def getTheServicePostFixNum():
#     pass




def gen_iplist(configList_,path):
    global serviceUrlIpListDict
    global serviceUrlIpListUserDict
    global del_serviceUrlIpListUserDict
    global configList
    global allEntryIdDict
    configList=configList_
    #用个全局变量来存储ip_prefix_list的最大值
    global ip_prefix_list_max_number
    ip_prefix_list_max_number = 50
    global log_list
    log_list = []
    serviceUrlIpListDict = {}
    serviceUrlIpListUserDict = {}
    del_serviceUrlIpListUserDict = {}
    global servicePortListDict
    servicePortListDict = {}
    global portListCommandList
    portListCommandList = []
    commandList = []
    ccl7_cfg = open(path + '\\configureL7.log', 'r')
    ccl7_cfg_list = ccl7_cfg.readlines()
    ccl7_cfg.close()
    allEntryIdList = eval(ccl7_cfg_list[0])
    serviceDict = eval(ccl7_cfg_list[1])
    serviceEntryIdDict = eval(ccl7_cfg_list[2])
    allEntryIdDict = eval(ccl7_cfg_list[3])
    servicePortListDict = eval(ccl7_cfg_list[4])

    excel_path = path+"\\ip_prefix_list_L7.xlsx"

    global servce_ipostfix_num
    servce_ipostfix_num = {}
    global servicePostFixNumDict
    servicePostFixNumDict = {}
    try:
        excel = openpyxl.load_workbook(excel_path)
    except Exception as err:
        print(err)

    sheet = excel["ip_prefix_list_add"]
    try:
        sheet_del = excel["ip_prefix_list_del"]
    except Exception as err:
        pass

    serviceList = []
    #该函数会根据分割符来把一条条目中包含port range的分成若干条
    serviceList = getServiceListByList(sheet, 1)

    resultList = []
    resultList = arrangeTheList(serviceList)

    resultList = arrangeTheList_2(resultList)
    for resultLine in resultList:
        if len(resultLine) == 0:
            resultList.remove(resultLine)

    for resultlst in resultList:
        addServiceUrlIpListToDict(resultlst)

    #print("用户iplist",serviceUrlIpListDict)

    #添加业务url列表（该业务对应的url所有的需要添加的IP）
    for resultlst in resultList:
        addServiceUrlIpListToUserDict(resultlst)
    #print("用户需要添加的iplist",serviceUrlIpListUserDict)

    #print("这是配置文件中相关业务的数据(添加前)")
    log_list.append("这是配置文件中相关业务的数据(添加前)\n")
    for sNameKey in serviceUrlIpListDict:
        log_list.append(sNameKey+"\n")
        #print(sNameKey)
        for urlKey in serviceUrlIpListDict[sNameKey]:
            log_list.append("  " + str(urlKey) + "\n")
            #print("  " + urlKey)
            for linelst in serviceUrlIpListDict[sNameKey][urlKey]:
                log_list.append("    " +str(len(linelst)-1)+"  "+str(linelst) + "\n")
                #print("    " +str(len(linelst)-1)+"  "+str(linelst))

    #先对业务进行增加操作
    #commandList.append("exit all\n")
    #commandList.append("configure application-assurance group 1:1 policy\n")
    for sNameKey in serviceUrlIpListUserDict:
        #commandList.append("对"+sNameKey+"业务进行新增操作\n")
        log_list.append("对"+sNameKey+"业务进行新增操作\n")
        for urlportKey in serviceUrlIpListUserDict[sNameKey]:
            log_list.append("对该业务的该URL进行添加:"+sNameKey+"  "+ str(urlportKey)+"  "+str(serviceUrlIpListUserDict[sNameKey][urlportKey]) + "\n")
            addCommandTocommandList(commandList,sNameKey, urlportKey,serviceUrlIpListUserDict[sNameKey][urlportKey])

    #print("这是配置文件中相关业务的数据(添加后)")
    for sNameKey in serviceUrlIpListDict:
        log_list.append(sNameKey + "\n")
        print(sNameKey)
       # commandList.append("对"+sNameKey+"业务进行添加"+"\n")
        for urlportKey in serviceUrlIpListDict[sNameKey]:
            log_list.append("  " + str(urlportKey) + "\n")
            commandList.append("   对" + str(urlportKey) + "进行操作" + "\n")
            #print("  " + urlKey)
            for linelst in serviceUrlIpListDict[sNameKey][urlportKey]:
                log_list.append("    " +str(len(linelst)-1)+str(linelst) + "\n")
                #print("    " +str(len(linelst)-1)+str(linelst))
                for line in linelst:
                    if "create" in line:
                        commandList.append('exit all\n')
                        commandList.append('configure application-assurance group 1:1\n')
                        commandList.append(line.replace(" create", "").replace("  ", "") + "\n\n")
                    elif "ip-prefix-list" in line and "create" not in line:
                        entryId = getTheCompatibleEntryIdByDict()
                        createIpPrefixListEntry(commandList, sNameKey, urlportKey, line, entryId)
                        commandList.append('exit all\n')
                        commandList.append('configure application-assurance group 1:1\n')
                        commandList.append(line.replace(" ", "") +" create" "\n\n")
                    elif "name" not in line:
                        addIpPrefixListCommand(commandList, sNameKey, line)


    fo = open(path+"\\脚本文件\\ip_prefix_list.txt", "w")
    fo.writelines(commandList)
    fo.close()

    text_cfg = []
    text_cfg.append(str(allEntryIdList) + "\n")
    text_cfg.append(str(serviceDict) + "\n")
    text_cfg.append(str(serviceEntryIdDict) + "\n")
    text_cfg.append(str(allEntryIdDict) + "\n")
    text_cfg.append(str(servicePortListDict) + "\n")

    file = open(path + "\\configureL7.log", "w")
    file.writelines(text_cfg)
    file.close()

    if log_list:
        fo_log = open("tmp\\ip_prefix_list_add.log", "w")
        fo_log.writelines(log_list)
        fo_log.close()

        fo_log = open(path + "\\ip_prefix_list_add.log", "w")
        fo_log.writelines(log_list)
        fo_log.close()
