# coding:utf-8
import os
import openpyxl
import ccL7
import ccl34
import ccl7_iplist
import ccl7_iplist_del, time, zipfile,ccl7_ip_prefix_list_HeaderEnrich,ccl7_HeaderEnrich


def getServiceListByList(sheet, startRow):
    global serviceCaseList
    global head_enrich_list
    changeLag_col = 3
    serviceId_col = 8
    serviceName_col = 10
    ipAddressL3_col = 13
    protocolNumber_col = 14
    portNumberL4_col = 15
    urlL7_col = 16
    HeaderEnrich_col = 19
    serviceCase_col = 9
    layerLag = ""
    # firstLineServiceId = sheet.cell(row=startRow, column=serviceId_col).value
    retList = []

    for rowNumber in range(startRow, sheet.max_row + 1):
        changeLag = sheet.cell(row=rowNumber, column=changeLag_col).value
        serviceId = sheet.cell(row=rowNumber, column=serviceId_col).value
        serviceName = sheet.cell(row=rowNumber, column=serviceName_col).value
        ipAddressL3 = sheet.cell(row=rowNumber, column=ipAddressL3_col).value
        protocolNumber = sheet.cell(row=rowNumber, column=protocolNumber_col).value
        portNumberL4 = sheet.cell(row=rowNumber, column=portNumberL4_col).value
        urlL7 = sheet.cell(row=rowNumber, column=urlL7_col).value
        if urlL7 != None and "." not in urlL7:
            urlL7 = None
        HeaderEnrich = sheet.cell(row=rowNumber, column=HeaderEnrich_col).value
        serviceCase = sheet.cell(row=rowNumber, column=serviceCase_col).value
        '''
        if "免" in serviceCase:
            serviceCase = "免"
        elif "统" in serviceCase:
            serviceCase = "统"
        elif "定" in serviceCase:
            serviceCase = "定"
        elif "收" in serviceCase:
            serviceCase = "收"
        else:
            serviceCase = "定"
        '''
        #print("6++++++++++6",str((serviceName,serviceCase)))
        serviceCaseList.append((serviceName,serviceCase))

        if HeaderEnrich != None:
            if "头增强" in HeaderEnrich:
                #HeaderEnrich = "头增强"
                head_enrich_list.append(
                    (changeLag, serviceId, serviceName, ipAddressL3, protocolNumber, portNumberL4, urlL7))
        retList.append((changeLag, str(serviceId), serviceName, ipAddressL3, protocolNumber, portNumberL4, urlL7))
    serviceCaseList = list(set(serviceCaseList))
    #print("7++++++++++7", str(serviceCaseList))
    #mtds_config = open("tmp//免统定收配置.txt", "w")
    #mtds_config.writelines(str(serviceCaseList))
    #mtds_config.close()
    return retList


def arrangeTheList(lst, configureList):
    sList = []
    retList = []
    tempList = []
    for tup in lst:
        changeLag, serviceId, serviceName, ipAddress, protocolNumber, portNumber, url  = tup
        sList.append(str(serviceId))
    #print(sList)
    sList = list(set(sList))
    #print("service id is", len(sList),sList)
    for sValue in sList:
        for tup in lst:
            changeLag, serviceId, serviceName, ipAddress, protocolNumber, portNumber, url = tup
            if sValue == serviceId:
                tempList.append(tup)
        retList.append(tempList)
        tempList = []
    #print("+++",len(sList),retList)
    newRetList = []
    tlst = []
    for retline in retList:
        layerLag = selectL347lag(retline, configureList)
        for tup in retline:
            changeLag, serviceId, serviceName, ipAddress, protocolNumber, portNumber, url = tup
            tlst.append((layerLag, changeLag, serviceId, serviceName, ipAddress, protocolNumber, portNumber, url))
        newRetList.append(tlst)
        tlst = []
    return newRetList


def selectL347lag(tupList, cfgLst):
    global log_list
    for i in range(0, len(cfgLst)):
        try:
            if 'policy-rule-unit "PRU_' + tupList[0][2] in cfgLst[i] and "qci * arp * precedence" not in cfgLst[i]:
                k = i
                for j in range(k, len(cfgLst)):
                    if "aa-charging-group" in cfgLst[j]:
                        #print("该业务" + tupList[0][2] + "是L7")
                        log_list.append("该业务" + tupList[0][2] + "是L7")
                        return "L7"
                    if "protocol" in cfgLst[j]:
                        #print("该业务" + tupList[0][2] + "是L4")
                        log_list.append("该业务" + tupList[0][2] + "是L4")
                        return "L4"
                    if "remote-ip" in cfgLst[j] and "protocol" not in cfgLst[j - 1]:
                        #print("该业务" + tupList[0][2] + "是L3")
                        log_list.append("该业务" + tupList[0][2] + "是L3")
                        return "L3"
                    if "exit" in cfgLst[j]:
                        break
        except:
            pass

    for tup in tupList:
        if tup[6] != None:
            #print("该业务", tupList[0][2] + "是L7为新业务")
            log_list.append("该业务"+tupList[0][2] + "是L7为新业务\n")
            return "L7"
    # print("++",tupList)
    for tup in tupList:
        if tup[4] != None or tup[5] != None:
            #print("该业务" + tupList[0][2] + "是L4为新业务")
            log_list.append("该业务" + tupList[0][2] + "是L4为新业务\n")
            return "L4"
    #print("该业务" + tupList[0][2] + "是L3为新业务")
    #print("99999",tupList)
    log_list.append("该业务" + tupList[0][2] + "是L3为新业务\n")
    return "L3"


def writeRowInExcel(sheet, px, py, writetup):
    sheet.cell(row=py, column=1, value=writetup[0])
    sheet.cell(row=py, column=3, value=writetup[1])
    sheet.cell(row=py, column=8, value=writetup[2])
    sheet.cell(row=py, column=10, value=writetup[3])
    sheet.cell(row=py, column=13, value=writetup[4])
    sheet.cell(row=py, column=14, value=writetup[5])
    sheet.cell(row=py, column=15, value=writetup[6])
    sheet.cell(row=py, column=16, value=writetup[7])


def getUrlTimes(tups7, urlstr):
    t = 0
    for line in tups7:
        if line[7] == urlstr:
            t += 1
    return t


def isIpList(serviceName, cfglist, http_host):
    http_host = http_host.replace("http://", "").replace("https://", "").replace("/*", "").replace(":*", "").split("/")[
        0]
    # print(serviceName,http_host)
    for i in range(0, len(cfglist)):
        if http_host in cfglist[i]:  #先判断http_host 再判断下面的server-address eq ....来确认是否是iplist
            k = i
            for j in range(k, len(cfglist)):
                if "no shutdown" in cfglist[j]:
                    break
                if 'server-address eq ip-prefix-list "app_' + serviceName in cfglist[j]:
                    return True
    return False


def getIPlistServiceTupList(tupLst7, configList):
    global log_list
    L7list = []
    urlList = []
    # url_times = []
    for tup in tupLst7:
        if tup[7] != None:
            L7list.append(tup)
            urlList.append(tup[7])
    # print(list(set(urlList)))
    urlList = list(set(urlList))
    serviceDict = {}
    iplist = []

    for tup in L7list:
        if isIpList(tup[3], configList, tup[7]) == True:
            iplist.append(tup)
    #print("66666+++++++",urlList)
    for url in urlList:
        if url != None:
            urlTime = getUrlTimes(tupLst7, url)
        # print(url + "出现次数:" + str(urlTime))

        log_list.append(url + "出现次数:" + str(urlTime) + "\n")
        # url_times.append(url + " APPEARS:" + str(urlTime)+" TIME(S)")
        if urlTime > 5:
            log_list.append("该" + url + "出现次数超过5次应放入ip-prefix-list:" + str(urlTime) + "\n")
            for tup in tupLst7:
                if tup[7] == url:
                    iplist.append(tup)

    # for line in iplist:
    #    print(line)
    return list(set(iplist))


def writeExcel(lst, postfix,configList, path):
    global log_list
    #print("++1",lst)
    tupListL34 = []
    #tupListL4 = []
    tupListL7 = []
    # url_times = []
    for llst in lst:
        if llst[0][0] == "L3":
            for tup in llst:
                tupListL34.append(tup)

    for llst in lst:
        if llst[0][0] == "L4":
            for tup in llst:
                tupListL34.append(tup)

    for llst in lst:
        if llst[0][0] == "L7":
            for tup in llst:
                tupListL7.append(tup)
    fPath = None
    #print("l34:",tupListL34)
    if len(tupListL34) != 0:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "L34"
        x = 1
        y = 3
        for tup in tupListL34:
            log_list.append(str(tup) + "该条目被存入‘内容计费整理L34’表格中" + "\n")
            writeRowInExcel(sheet, x, y, tup)
            y += 1
        if postfix ==None:
            fPath = path + "\\内容计费整理L34" + ".xlsx"
        else:
            fPath = path + "\\内容计费整理L34" +postfix+ ".xlsx"
        if fPath != None:
            wb.save(fPath)
            wb.close()

    fPath = None
    ip_list_tupList = []
    #print("l7:", tupListL7)
    if len(tupListL7) != 0:
        ip_list_tupList = getIPlistServiceTupList(tupListL7, configList)
        # url_times = getIPlistServiceTupList(tupListL7, configList)[1]
        # print("iplist+++",ip_list_tupList)
        # print("+++++++++++")
        for line in ip_list_tupList:
            # print(line)
            tupListL7.remove(line)

    if len(tupListL7) != 0:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "L7"
        x = 1
        y = 3
        for tup in tupListL7:
            log_list.append(str(tup) + "该条目被存入‘内容计费整理L7’表格中" + "\n")
            writeRowInExcel(sheet, x, y, tup)
            y += 1
        if postfix == None:
            fPath = path + "\\内容计费整理L7" + ".xlsx"
        else:
            fPath = path + "\\内容计费整理L7" +postfix+ ".xlsx"
        if fPath != None:
            wb.save(fPath)
            wb.close()
    fPath = None
    # print("+++",ip_list_tupList)
    del_ip_list_tupList = []
    add_ip_list_tupList = []
    for tup in ip_list_tupList:
        if tup[1] == "删除":
            del_ip_list_tupList.append(tup)
        if tup[1] == "新增":
            add_ip_list_tupList.append(tup)

    if len(add_ip_list_tupList) != 0 or len(del_ip_list_tupList) != 0:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "ip_prefix_list_add"
        sheet_del = wb.create_sheet("ip_prefix_list_del")
    #print("l7list_add:", add_ip_list_tupList)
    if len(add_ip_list_tupList) != 0:
        # wb = openpyxl.Workbook()
        # sheet = wb.active
        # sheet.title = "ip_prefix_list"
        x = 1
        y = 3
        for tup in add_ip_list_tupList:
            log_list.append(str(tup) + "该条目被存入‘ip_prefix_list_L7EXCEL的ip_prefix_list_add’表格中" + "\n")
            writeRowInExcel(sheet, x, y, tup)
            y += 1
        if postfix == None:
            fPath = path + "\\ip_prefix_list_L7" + ".xlsx"
        else:
            fPath = path + "\\ip_prefix_list_L7" +postfix+ ".xlsx"
        # wb.save(fPath)
        # wb.close()
    #print("l7list_del", del_ip_list_tupList)
    if len(del_ip_list_tupList) != 0:
        # del sheet
        # wb = openpyxl.Workbook()
        # sheet = wb.active
        # sheet.title = "ip_prefix_list_del"
        x = 1
        y = 3
        for tup in del_ip_list_tupList:
            log_list.append(str(tup) + "该条目被存入‘ip_prefix_list_L7EXCEL的ip_prefix_list_del’表格中" + "\n")
            writeRowInExcel(sheet_del, x, y, tup)
            y += 1
        if postfix == None:
            fPath =path + "\\ip_prefix_list_L7" + ".xlsx"
        else:
            fPath = path + "\\ip_prefix_list_L7" +postfix+ ".xlsx"
    try:
        if fPath !=None:
            wb.save(fPath)
            wb.close()
    except:
        pass
    return tupListL34, tupListL7, del_ip_list_tupList, add_ip_list_tupList


def gen_origin_api(*args):
    serviceDi = []
    path = ''
    global log_list
    log_list = []
    path = os.path.abspath('.')
    configFile = open(args[1], 'r')
    configList = configFile.readlines()
    configFile.close()

    global head_enrich_list
    head_enrich_list = []
    global serviceCaseList
    serviceCaseList = []

    for ne_name in configList:
        if ne_name.find('BNK"') != -1:
            ne_name = ne_name[ne_name.find('name "') + 6:-2]
            l_time = time.strftime('%Y%m%d_%H%M%S', time.localtime(time.time()))
            #print(args[0])
            path = path + '\\' + 'Generated\\' + ne_name + '\\' + l_time + '\\' + args[0][0:-5].replace('tmp\\','')
            mkdir(path)
            break
    excel = openpyxl.load_workbook(args[0])
    sheet = excel["本次变更"]
    serviceList = []
    serviceList = getServiceListByList(sheet, 3)

    resultList = arrangeTheList(serviceList, configList)
    resultList_head = arrangeTheList(head_enrich_list, configList)
    statistics_list = writeExcel(resultList,None,configList, path)

    if os.path.exists(path + "\\内容计费整理L34.xlsx"):
        ccl34.gen_l34(configList, path)
    # else:
    #     fo_log = open("L34.log", "w")
    #     fo_log.writelines(['本次无34层数据变更'])
    #     fo_log.close()
    if os.path.exists(path + "\\内容计费整理L7.xlsx"):
        serviceDi = ccL7.gen_l7(configList, path)
    # else:
    #     fo_log = open("L7.log", "w")
    #     fo_log.writelines(['本次无7层数据变更'])
    #     fo_log.close()
    if os.path.exists(path + "\\ip_prefix_list_L7.xlsx"):
        ccl7_iplist.gen_iplist(configList, path)
        ccl7_iplist_del.gen_iplist_del(configList, path)
    if resultList_head:
        writeExcel(resultList_head, "_headEnrich", configList, path)
        ccl7_HeaderEnrich.gen_hearderenrich(path,configList)
    if os.path.exists(path+"\\ip_prefix_list_L7_headEnrich.xlsx"):
        ccl7_ip_prefix_list_HeaderEnrich.gen_prefix_enrich(path,configList)


    fo = open(path + "\\processL347.log", "w")
    fo.writelines(log_list)
    fo.close()

    zipfile = path + '.zip'
    loc_fo = open("tmp\\processL347.log", "w")
    loc_fo.writelines(log_list)
    loc_fo.close()
    zip_ya(path, zipfile)
    # return zipfile


def mkdir(path):
    path = path.strip()
    # path = path.rstrip("\\")
    isExists = os.path.exists(path)
    if not isExists:
        os.makedirs(path)
        return True
    else:
        # print(path + ' 目录已存在,将直接覆盖旧文件...')
        return False


def zip_ya(startdir, file_news):
    z = zipfile.ZipFile(file_news, 'w', zipfile.ZIP_DEFLATED)  # 参数一：文件夹名
    for dirpath, dirnames, filenames in os.walk(startdir):
        fpath = dirpath.replace(startdir, '')  # 这一句很重要，不replace的话，就从根目录开始复制
        fpath = fpath and fpath + os.sep or ''  # 这句话理解我也点郁闷，实现当前文件夹以及包含的所有文件的压缩
        for filename in filenames:
            z.write(os.path.join(dirpath, filename), fpath + filename)
            # print('压缩成功')
    z.close()
